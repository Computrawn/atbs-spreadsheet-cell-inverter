#! python3
# spreadsheetCellInverter.py — An exercise in manipulating Excel files.
# For more information, see project_details.txt.

import openpyxl
from openpyxl.utils import get_column_letter

file_name = input("Please enter filename here: ")
xy_list = []


def cell_inverter(document):
    """Opens xlsx file, casts the cells into list of lists data structure
    and writes the inverted data to a new worrkbook."""
    wb = openpyxl.load_workbook(f"{document}.xlsx")
    sheet = wb.active

    for column in range(1, sheet.max_column + 1):
        column_letter = get_column_letter(column)
        column_list = []
        for row in range(1, sheet.max_row + 1):
            column_list.append(sheet[f"{column_letter}{row}"].value)
        xy_list.append(column_list)

    inv_wb = openpyxl.Workbook()
    inv_sheet = inv_wb.active
    for y in range(1, sheet.max_column + 1):
        for x in range(1, sheet.max_row + 1):
            letter = get_column_letter(x)
            inv_sheet[f"{letter}{y}"] = xy_list[y - 1][x - 1]

    inv_wb.save(f"{document}_inverted.xlsx")


cell_inverter(file_name)
