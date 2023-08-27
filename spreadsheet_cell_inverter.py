#!/usr/bin/env python3
# spreadsheetCellInverter.py — An exercise in manipulating Excel files.
# For more information, see README.md

import logging
import openpyxl
from openpyxl.utils import get_column_letter

logging.basicConfig(
    level=logging.DEBUG,
    filename="logging.txt",
    format="%(asctime)s -  %(levelname)s -  %(message)s",
)
logging.disable(logging.CRITICAL)  # Note out to enable logging.


def cell_inverter(document: str) -> None:
    """Opens xlsx file, casts the cells into list of lists data structure
    and writes the inverted data to a new workbook."""

    wb = openpyxl.load_workbook(f"{document}.xlsx")
    sheet = wb.active
    xy_list = [
        [
            sheet[f"{get_column_letter(column)}{row}"].value
            for row in range(1, sheet.max_row + 1)
        ]
        for column in range(1, sheet.max_column + 1)
    ]

    inv_wb = openpyxl.Workbook()
    inv_sheet = inv_wb.active

    for y_value in range(1, sheet.max_column + 1):
        for x_value in range(1, sheet.max_row + 1):
            inv_sheet[f"{get_column_letter(x_value)}{y_value}"] = xy_list[y_value - 1][
                x_value - 1
            ]

    inv_wb.save(f"{document}_inverted.xlsx")


def main() -> None:
    cell_inverter(input("Please enter filename here: "))


if __name__ == "__main__":
    main()
